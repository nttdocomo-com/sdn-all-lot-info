#特定のフォルダに格納されたファイル一覧を取得するためのモジュール
import glob
import datetime

#CSV形式のファイルを扱うためのモジュール
import csv

#Excel形式のファイルを扱うためのモジュール
import openpyxl as excel
from openpyxl.styles import Font, PatternFill, Border, Side


#地域エリアを変換するための関数
def load_conversion_map(filename):
    conversion_map = {}
    with open(filename, mode='r', encoding='utf-8-sig') as f:
        reader = csv.reader(f) 
        headers=next(reader)
        #print(f"Headers: {headers}")
        for row in reader:
            #print(row)
            conversion_map[row[0]] = row[1]  
    return conversion_map

#ビル名称を変換するための関数
def load_building_map(filename):
    conversion_map = {}
    with open(filename, mode='r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        headers=next(reader)
        #print(f"Headers: {headers}")
        for row in reader:
            #print(row)
            conversion_map[row['buil']] = row['ビル']  
    return conversion_map

#上記変換関数を用いてロット名表記を変換する関数
def convert_lot_name(lot_name, area_map, building_map):
    area_id, building_id, suffix = extract_lot_identifier(lot_name)
    
    # area_idとbuilding_idをマップで変換
    area_name = area_map.get(area_id, area_id)  # area_mapから変換
    building_name = building_map.get(building_id, building_id)  # building_mapから変換
    
    # 最終的なシート名を作成する。末尾を大文字に変換して追加。
    return f"{area_name}_{building_name}_{suffix.upper()}"

#ロット名変換のためにハイフン区切りを解消する関数
def extract_lot_identifier(lot_name):
    parts = lot_name.split('-')
    if len(parts) >= 3:
        return parts[0],parts[1],parts[2]  
    return lot_name,None ,None

def arrange():
    print("arrange処理を開始します")

    # 変換マップをファイルから読み込み
    # 変換表は実行ファイルと同ディレクトリに格納　※新設に伴いメンテ要
    area_map = load_conversion_map(".\const\conversion-area.csv")
    building_map = load_conversion_map(".\const\conversion-building.csv")

    # ログファイル一覧を取得
    files = glob.glob(".\silent-check\*all_lot.csv")
    files.sort()
    #print(files) #for-debug
    # 最新ファイルにのみ処理を実施
    if files:
        latest_file = files[-1]
        # print(latest_file) #for-debug
    else:
        print("対象となるファイルが存在しません")
        return  # 処理を終了

    # データを保持するリスト
    records = []
    lot_array = []

    if latest_file:
        with open(latest_file) as f:
            reader = csv.reader(f)
            for row in reader:
                if row[0].strip() != "lot":  # ヘッダー行をスキップ
                    lot = row[0]
                    lot_array.append(lot)
                    records.append(row)  # 各行を記録

        # 取得したロット名をソートして重複排除
        lot_array.sort()
        lot_array = list(set(lot_array))

    #lot名を格納するリスト
    sorted_lot_names = sorted(lot_array, key =lambda lot: convert_lot_name(lot,area_map,building_map))
    # print("sort_lot_names示します") #for-debug
    # print(sorted_lot_names) #for-debug

    # Excelファイルの作成
    wb = excel.Workbook()

    # デフォルトで存在する不要なシートを削除
    wb.remove(wb['Sheet'])


    # lot毎にシートを作成
    for lot_name in sorted_lot_names:
        #変換後の形式でシート名を設定
        converted_name = convert_lot_name(lot_name, area_map, building_map)
        #print(converted_name) #for-debug
        ws = wb.create_sheet(title=converted_name)

        #ヘッダ定義
        headers = ["ID", "Name", "Serial Number", "Verion","PID"]
        ws.append(headers) #ヘッダ情報の書き込み

        #ヘッダスタイル設定
        header_font=Font(bold=True)
        header_fill=PatternFill(start_color="00FF00", end_color="00FF00",fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill  

        # データを追加
        for row in records:
            if row[0] == sorted_lot_names:
                # print(row[0])
                # print("row-lot_name")
                # print(lot_name) #for-debug
                data_to_append = row[1],row[5],row[4],row[2],row[6] #ID,Name,S/N,OSVer,PID の順に抽出
                ws.append(data_to_append)

                #罫線を引くためのスタイル
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                
                #格子罫線を引く
                for cell in ws[f'A{ws.max_row}':f'E{ws.max_row}'][0]:
                    cell.border = thin_border

        # セル幅の自動調節
        for column in ws.columns:
            max_length = 0
            col_letter = excel.utils.get_column_letter(column[0].column)  # 列のアルファベットを取得
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 5  # 幅を調整（+5は余白）
            ws.column_dimensions[col_letter].width = adjusted_width  # 幅を設定

    # Excelファイルを保存
    logfile = 'all-lot-info.xlsx'
    # wb.save(logfile)

    print("処理完了")


    #####################################



    ####共用PCで起動する際は以下のコメントアウトを外す#####


    #日時モジュールをインポート
    import datetime
    dt_now = datetime.datetime.now() #2022/10/26 追加
    day = dt_now.strftime('%Y-%m-%d-') #2022/10/26 追加
    logfile = day + 'all_lot_info.xlsx' #2022/10/26 追加

    #Excelファイルの保存
    wb.save(logfile)
    print("Excelを保存 ： " + logfile)

    #ファイル移動などのモジュールインポート
    import shutil
    #ファイルを移動
    try:
        shutil.move(logfile, r'./testbox/')
        # print("success") #for-debug
    except:
       #エラ―解析用   
       print("error")   #for-debug
       import traceback
       with open('silent_kenchi.log', 'a') as f:
            traceback.print_exc(file=f)

    ###########################################
arrange()